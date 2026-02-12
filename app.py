# app.py
from __future__ import annotations

import re
import uuid
import copy

from pathlib import Path
from typing import Tuple, List, Dict, Any, Optional

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO

from datetime import datetime, timedelta

import psycopg2
from db_cache import (
    cached_load_stocks,
    cached_load_templates,
    cached_load_template_items,
    cached_load_plans,
    cached_eps_db,
    cached_cf_db,
    cached_reactions_in_plan,
    cached_labeling_records
)

from utils import (
    slugify, fmt_num, auto_stock_id,
    conc_from_amount_volume,
    to_mM, from_mM,
    to_uL, from_uL,
    amount_nmol_from_conc_vol,
    calc_volume_uL_from_target,
    compute_reaction,
    lookup_cf
)

from db_utils_pg import (
    get_eps_db, execute,

    load_stocks, insert_stock, update_stock, delete_stock,
    load_plans, 

    load_templates, load_template_items, delete_template, update_template_meta,
    save_template_from_computed,

    upsert_epsilon, delete_epsilon,
    get_epsilon_value,
    
    get_cf, upsert_cf, get_cf_db,delete_cf,
    fetchall
)

CONC_UNITS = ["M", "mM", "uM", "nM"]
VOL_UNITS = ["uL", "mL"]
AMT_UNITS = ["nmol", "pmol", "umol"]

CATEGORIES = [
    "Buffer preparation",
    "Chemical reaction",
    "Cell dish",
    "Labeling",
    "Other",
]


# render
def render_template_manager(conn):
    st.subheader("Template Manager")

    tmpl_df = cached_load_templates()
    if "creating_template" not in st.session_state:
        st.session_state["creating_template"] = False
    if "editing_template" not in st.session_state:
        st.session_state["editing_template"] = False
    if "tmpl_rows" not in st.session_state:
        st.session_state["tmpl_rows"] = None

    # ---- Add new template ----
    st.markdown("### Template")

    tmpl_names = ["➕ New template"] + (
        tmpl_df["name"].tolist() if not tmpl_df.empty else []
    )

    sel_name = st.selectbox(
        "Select template",
        tmpl_names,
        key="template_select_box"
    )

    if sel_name == "➕ New template":
        creating = True
        row = None
        template_id = None
    else:
        creating = False
        row = tmpl_df[tmpl_df["name"] == sel_name].iloc[0]
        template_id = row["template_id"]

    st.session_state["creating_template"] = creating

    st.divider()

    # ---- Template meta 수정 ----
    st.markdown("### Template info")

    if st.session_state["creating_template"]:
        new_name = st.text_input(
            "Template name *",
            value=st.session_state.get("new_template_name", "")
        )
        new_desc = st.text_input(
            "Description",
            value=st.session_state.get("new_template_desc", "")
        )
    else:
        new_name = st.text_input(
            "Template name",
            value=row["name"]
        )
        new_desc = st.text_input(
            "Description",
            value=row.get("description") or ""
        )


    col1, col2 = st.columns([1, 1])

    if col1.button("Update template info"):
        if not new_name.strip():
            st.error("Template name은 비워둘 수 없습니다.")
        else:
            update_template_meta(conn, template_id, new_name.strip(), new_desc)
            cached_load_templates.clear()
            st.success("Template info updated.")
            st.rerun()
            

    if col2.button("Delete template"):
        st.warning("이 작업은 되돌릴 수 없습니다.")
        if st.button("정말 삭제할까요?"):
            delete_template(conn, template_id)
            cached_load_templates.clear()
            st.success("Template deleted.")
            st.rerun()
            

    st.divider()
    st.markdown("### Final volume")
    if st.session_state["creating_template"]:
        tmpl_final_volume = st.number_input(
            "Final volume",
            min_value=0.0,
            value=st.session_state.get("tmpl_final_volume", 20.0),
            step=1.0,
            key="tmpl_final_volume"
        )
        tmpl_final_vol_unit = st.selectbox(
            "Unit",
            VOL_UNITS,
            index=VOL_UNITS.index(
                st.session_state.get("tmpl_final_vol_unit", "uL")
            ),
            key="tmpl_final_vol_unit"
        )
    else:
        tmpl_final_volume = st.number_input(
            "Final volume",
            min_value=0.0,
            value=float(row.get("final_volume", 20.0)),
            step=1.0,
            key="tmpl_final_volume"
        )
        tmpl_final_vol_unit = st.selectbox(
            "Unit",
            VOL_UNITS,
            index=VOL_UNITS.index(
                row.get("final_volume_unit", "uL")
            ),
            key="tmpl_final_vol_unit"
        )

    # ---- Template components ----
    st.markdown("### Template components")

    if template_id:
        items = cached_load_template_items(template_id)
    else:
        items = pd.DataFrame()


    # 편집 모드 상태
    if "editing_template" not in st.session_state:
        st.session_state["editing_template"] = False

    if "tmpl_rows" not in st.session_state:
        st.session_state["tmpl_rows"] = None


    
    # READ-ONLY MODE
    if not st.session_state["editing_template"]:
        if items.empty:
            st.info("Template에 저장된 component가 없습니다.")
        else:
            st.dataframe(
                items.rename(columns={
                    "stock_id": "Stock ID",
                    "example_target": "Target conc",
                    "example_target_unit": "Target unit",
                    "example_volume": "Volume",
                    "example_volume_unit": "Volume unit",
                    "item_note": "Note"
                }),
                width='stretch',
                hide_index=True
            )

        if st.button("✏️ Update template"):
            st.session_state["editing_template"] = True
            st.session_state["tmpl_rows"] = None
            st.rerun()

    # EDIT MODE (reaction card
    else:
        st.caption(
            "This editor allows calculation for template validation only. "
            "Actual experiment records are created in New Reaction."
        )

        stocks = cached_load_stocks()
        stock_options = ["(DB 미등록: 임시 시약)"] + stocks["id"].tolist()

        # 최초 진입 시 template_items → rows 변환
        if st.session_state["tmpl_rows"] is None:
            rows = []
            if not items.empty:
                for _, r in items.iterrows():
                    rows.append({
                        "row_id": str(uuid.uuid4()),
                        "stock_sel": r["stock_id"] or "(DB 미등록: 임시 시약)",
                        "custom_name": "",
                        "custom_stock_conc": 0.0,
                        "custom_stock_unit": "mM",
                        "target_conc": float(r["example_target"] or 0.0),
                        "target_unit": r["example_target_unit"] or "mM",
                        "vol": float(r["example_volume"] or 0.0),
                        "vol_unit": r["example_volume_unit"] or "uL",
                        "note": r["item_note"] or "",
                    })
            else:
                rows = [empty_row()]

            st.session_state["tmpl_rows"] = rows

        # ---- Add row ----
        if st.button("+ Add reagent"):
            st.session_state["tmpl_rows"].append(empty_row())
            st.rerun()

        # ---- Render rows (reaction card style) ----
        for i, row in enumerate(st.session_state["tmpl_rows"]):
            with st.container(border=True):
                c1, c2, c3, c4, c5, c6, c7 = st.columns([2.2,1.2,1,1.2,1,2.2,0.6])
                rid = row["row_id"]

                row["stock_sel"] = c1.selectbox(
                    "Stock",
                    stock_options,
                    index=stock_options.index(row["stock_sel"]) if row["stock_sel"] in stock_options else 0,
                    key=f"tmpl_stock_{rid}"
                )

                row["target_conc"] = c2.number_input(
                    "Target conc",
                    min_value=0.0,
                    step=0.0001,
                    value=row["target_conc"],
                    key=f"tmpl_tc_{rid}"
                )

                row["target_unit"] = c3.selectbox(
                    "Unit",
                    CONC_UNITS,
                    index=CONC_UNITS.index(row["target_unit"]),
                    key=f"tmpl_tu_{rid}"
                )

                row["vol"] = c4.number_input(
                    "Volume",
                    min_value=0.0,
                    value=row["vol"],
                    key=f"tmpl_vol_{rid}"
                )

                row["vol_unit"] = c5.selectbox(
                    "Vol unit",
                    VOL_UNITS,
                    index=VOL_UNITS.index(row["vol_unit"]),
                    key=f"tmpl_vu_{rid}"
                )

                row["note"] = c6.text_input(
                    "Note",
                    row["note"],
                    key=f"tmpl_note_{rid}"
                )

                if c7.button("🗑", key=f"tmpl_del_{rid}"):
                    st.session_state["tmpl_rows"].pop(i)
                    st.rerun()

        
        # Calculation preview
        
        st.divider()
        st.markdown("### Calculation preview")

        tmpl_card = {
            "name": "Template preview",
            "final_volume": tmpl_final_volume,
            "final_vol_unit": tmpl_final_vol_unit,
            "rows": st.session_state["tmpl_rows"],
        }


        errors, computed, total_uL, final_uL_total = compute_reaction(
            tmpl_card, stocks, stock_options
        )

        if errors:
            st.error("Validation failed:")
            for e in errors:
                st.write(f"- {e}")
        else:
            if computed:
                df = pd.DataFrame(computed).sort_values("line_no")
                st.dataframe(
                    df[
                        [
                            "line_no",
                            "reagent",
                            "stock_conc",
                            "stock_unit",
                            "target_conc",
                            "target_unit",
                            "volume",
                            "volume_unit",
                            "amount",
                            "amount_unit",
                            "note",
                        ]
                    ],
                    width='stretch',
                    hide_index=True
                )

                fill_uL = final_uL_total - total_uL
                st.success(f"Fill remaining: {fill_uL:.3f} uL")
            else:
                st.info("No computable rows yet.")

        
        # Save / Cancel
        c1, c2 = st.columns(2)

        # -----------------------------
        # CREATE NEW TEMPLATE
        # -----------------------------
        if st.session_state["creating_template"]:
            if c1.button("💾 Create template"):
                if not new_name.strip():
                    st.error("Template name is required.")
                    st.stop()

                # 1️⃣ template 먼저 생성
                new_template_id = str(uuid.uuid4())

                execute(
                    """
                    INSERT INTO templates (
                        template_id, name, description,
                        final_volume, final_volume_unit
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        new_template_id,
                        new_name.strip(),
                        new_desc.strip() or None,
                        tmpl_final_volume,
                        tmpl_final_vol_unit,
                    )
                )

                # 2️⃣ template_items 저장
                for i, r in enumerate(st.session_state["tmpl_rows"], start=1):
                    execute(
                        """
                        INSERT INTO template_items (
                            template_id, line_no,
                            stock_id,
                            example_target, example_target_unit,
                            example_volume, example_volume_unit,
                            item_note
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            new_template_id,
                            i,
                            None if r["stock_sel"] == "(DB 미등록: 임시 시약)" else r["stock_sel"],
                            r["target_conc"] or None,
                            r["target_unit"],
                            r["vol"] or None,
                            r["vol_unit"],
                            r["note"] or None,
                        )
                    )

                cached_load_templates.clear()
                cached_load_template_items.clear()

                st.session_state["creating_template"] = False
                st.session_state["editing_template"] = False
                st.session_state["tmpl_rows"] = None
                
                if "template_select_box" in st.session_state:
                    del st.session_state["template_select_box"]
                st.success("New template created.")
                st.rerun()

        # -----------------------------
        # UPDATE EXISTING TEMPLATE
        # -----------------------------
        else:
            if c1.button("💾 Save changes"):
                execute("DELETE FROM template_items WHERE template_id = %s", (template_id,))

                for i, r in enumerate(st.session_state["tmpl_rows"], start=1):
                    execute(
                        """
                        INSERT INTO template_items (
                            template_id, line_no,
                            stock_id,
                            example_target, example_target_unit,
                            example_volume, example_volume_unit,
                            item_note
                        )
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            template_id,
                            i,
                            None if r["stock_sel"] == "(DB 미등록: 임시 시약)" else r["stock_sel"],
                            r["target_conc"] or None,
                            r["target_unit"],
                            r["vol"] or None,
                            r["vol_unit"],
                            r["note"] or None,
                        )
                    )

                # meta update
                update_template_meta(
                    conn,
                    template_id,
                    new_name.strip(),
                    new_desc.strip(),
                    tmpl_final_volume,
                    tmpl_final_vol_unit,
                )

                cached_load_template_items.clear()
                cached_load_templates.clear()

                st.success("Template updated.")
                st.rerun()


    
# ---------------- reaction card state ----------------
def empty_row():
    return {
        "row_id": str(uuid.uuid4()),  # 🔑 핵심
        "stock_sel": "(DB 미등록: 임시 시약)",
        "custom_name": "",
        "custom_stock_conc": 0.0,
        "custom_stock_unit": "mM",
        "target_conc": 0.0,
        "target_unit": "mM",
        "vol": 0.0,
        "vol_unit": "uL",
        "note": "",
    }



def new_reaction_card(idx: int):
    return {
        "name": f"Reaction {idx + 1}",
        "final_volume": 20.0,
        "final_vol_unit": "uL",
        "rows": [empty_row()],
        "template_select": "(None)",
        "template_name_for_save": "",
        "template_desc_for_save": "",
        "include_in_save_all": True,
        "fv_key_ver": 0
    }


# ---------------- save all cards ----------------
def save_all_to_db(
    conn,
    plan_id: Optional[str],
    plan_title: str,
    plan_notes: str,
    category: str,
    username: str,
    title_prefix: str,
    cards_to_save: List[Tuple[Dict[str, Any], List[Dict[str, Any]]]],
) -> Tuple[bool, str, str]:
    """
    cards_to_save: list of (card, computed)
    returns (ok, message, plan_id)
    """
    if len(cards_to_save) == 0:
        return False, "저장할 reaction이 없습니다. (Include 체크 확인)", plan_id or ""

    # Create plan if needed
    pid = plan_id
    if not pid:
        pid = str(uuid.uuid4())
        plan_title_clean = (plan_title or "").strip() or f"{pd.Timestamp.now():%Y%m%d}_{username}"
        execute(
            "INSERT INTO plans(plan_id, title, category, created_by, notes) VALUES (%s, %s, %s, %s, %s)",
            (pid, plan_title_clean, category, username, plan_notes or None)
        )

    # Transaction
    try:
        for idx, (card, computed) in enumerate(cards_to_save, start=1):
            reaction_id = str(uuid.uuid4())
            rx_name = (card.get("name") or f"Reaction {idx}").strip()
            rx_title = (title_prefix or "").strip() or ""
            if rx_title:
                full_title = f"{rx_title} :: {rx_name}"
            else:
                full_title = rx_name

            final_volume = float(card.get("final_volume") or 0.0)
            final_vol_unit = str(card.get("final_vol_unit") or "uL")

            execute(
                """
                INSERT INTO reactions(
                reaction_id, title, category, created_by, final_volume, final_volume_unit, plan_id
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (reaction_id, full_title, category, username, final_volume, final_vol_unit, pid)
            )

            for it in computed:
                execute(
                    """
                    INSERT INTO reaction_items(
                    reaction_id, line_no, stock_id, custom_name, stock_conc, stock_unit,
                    target_conc, target_conc_unit, volume, volume_unit, amount, amount_unit, note
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        reaction_id, int(it["line_no"]), it.get("stock_id"), it.get("custom_name"),
                        float(it.get("stock_conc") or 0.0), it.get("stock_unit"),
                        float(it.get("target_conc") or 0.0), it.get("target_unit"),
                        float(it.get("volume") or 0.0), it.get("volume_unit"),
                        float(it.get("amount") or 0.0), it.get("amount_unit") or "nmol",
                        it.get("note") or None
                    )
                )

            
        return True, f"Saved {len(cards_to_save)} reaction(s) into plan_id={pid}", pid
    except Exception as e:
        # conn.rollback()
        return False, f"DB 저장 실패: {e}", pid or ""

# export
def export_reactions_to_excel(
    reactions: List[Tuple[str, pd.DataFrame]]
) -> BytesIO:
    """
    reactions: [(reaction_name, result_df), ...]
    result_df columns expected:
    [
        line_no, reagent,
        stock_conc, stock_unit,
        target_conc, target_unit,
        volume, volume_unit,
        amount, amount_unit,
        note
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reactions"

    row_cursor = 1
    header_font = Font(bold=True)

    for rx_name, df in reactions:
        # -----------------------------
        # Reaction title
        # -----------------------------
        ws.cell(row=row_cursor, column=1, value=f"Reaction: {rx_name}")
        ws.cell(row=row_cursor, column=1).font = Font(bold=True)
        row_cursor += 2

        # -----------------------------
        # 🔑 Export 전용 DF 가공
        # -----------------------------
        df_export = df.copy()

        # 값 + unit 합치기
        df_export["Stock conc"] = (
            df_export["stock_conc"].astype(str)
            + " "
            + df_export["stock_unit"]
        )

        df_export["Target conc"] = (
            df_export["target_conc"].astype(str)
            + " "
            + df_export["target_unit"]
        )

        df_export["Volume"] = (
            df_export["volume"].astype(str)
            + " "
            + df_export["volume_unit"]
        )

        df_export["Amount"] = (
            df_export["amount"].astype(str)
            + " "
            + df_export["amount_unit"]
        )

        # 엑셀에 쓸 컬럼만 선택 (사람 기준)
        df_export = df_export[
            [
                "line_no",
                "reagent",
                "Stock conc",
                "Target conc",
                "Volume",
                "Amount",
                "note",
            ]
        ]

        # -----------------------------
        # Header
        # -----------------------------
        for col_idx, col in enumerate(df_export.columns, start=1):
            cell = ws.cell(row=row_cursor, column=col_idx, value=col)
            cell.font = header_font

        row_cursor += 1

        # -----------------------------
        # Table body
        # -----------------------------
        for _, r in df_export.iterrows():
            for col_idx, col in enumerate(df_export.columns, start=1):
                ws.cell(row=row_cursor, column=col_idx, value=r[col])
            row_cursor += 1

        row_cursor += 2  # blank lines between reactions

    # -----------------------------
    # Auto column width
    # -----------------------------
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------- main ----------------
def main():
    st.set_page_config(page_title="LabCalc", layout="wide")
    conn = None
    e_page=None


    st.title("LabCalc")

    with st.sidebar:
        st.header("User")
        username = st.text_input("Name", value="JW")
        st.divider()
        st.header("Reaction Menu")
        page = st.radio(
            "Reaction plan helper",
            ["New Reaction", "Plans", "Templates", "Stock DB", "Nanodrop"],
            index=0
        )

        if page == "Nanodrop":
            st.header("Efficiency")

            e_page = st.radio(
                "Labeling / UV-Vis tools",
                ["Labeling Efficiency", "Labeling Records", "Epsilon DB"],
                index=0
            )


            st.divider()


    # ---------------- Stock DB ----------------
    if page == "Stock DB":
        st.subheader("Stock DB (등록/수정/삭제)")

        stocks = cached_load_stocks()
        st.dataframe(stocks, width='stretch', hide_index=True,
                     column_config={
                    "notes": st.column_config.TextColumn(
                        "notes", width="large"
                    )}
                    )

        st.divider()
        st.markdown("### Add / Update Stock")

        mode = st.radio("Mode", ["Add", "Update", "Delete"], horizontal=True)

        if mode == "Add":
            with st.form("add_stock"):
                name = st.text_input("Reagent name *")
                notes = st.text_input("Notes (optional)")
                st.markdown("**입력 방식 선택**")
                input_mode = st.radio("Choose input", ["Concentration", "Amount + Volume"], horizontal=True)

                if input_mode == "Concentration":
                    c1, c2 = st.columns([1, 1])
                    with c1:
                        stock_conc = st.number_input("Stock concentration *", min_value=0.0, value=0.0, step=0.0001, format="%.4f")
                    with c2:
                        stock_unit = st.selectbox("Unit *", CONC_UNITS, index=0)
                else:
                    a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
                    with a1:
                        amount = st.number_input("Amount *", min_value=0.0, value=0.0, step=0.01)
                    with a2:
                        amount_unit = st.selectbox("Amount unit *", AMT_UNITS, index=0)
                    with a3:
                        volume = st.number_input("Volume *", min_value=0.0, value=0.0, step=0.01)
                    with a4:
                        volume_unit = st.selectbox("Volume unit *", VOL_UNITS, index=0)

                    stock_conc_mM, stock_unit = conc_from_amount_volume(amount, amount_unit, volume, volume_unit)
                    stock_conc = float(stock_conc_mM)
                    st.info(f"Calculated concentration: {fmt_num(stock_conc)} mM")

                custom_id = st.text_input("Stock ID (optional). 비우면 자동 생성", value="")

                submitted = st.form_submit_button("Save")
                if submitted:
                    if not name.strip():
                        st.error("Reagent name은 필수입니다.")
                        st.stop()

                    if stock_conc <= 0:
                        st.error("Stock concentration은 0보다 커야 합니다.")
                        st.stop()

                    stock_id = custom_id.strip() or auto_stock_id(name, stock_conc, stock_unit)

                    try:
                        insert_stock(conn, stock_id, name.strip(), stock_conc, stock_unit, notes.strip())
                        st.success(f"Saved: {stock_id}")
                        cached_load_stocks.clear()
                        
                    except psycopg2.IntegrityError as e:
                        st.error("중복이거나(id 또는 name+conc+unit), 제약조건 위반입니다.")
                        cached_load_stocks.clear()
                        st.caption(str(e))

        elif mode == "Update":
            stocks = cached_load_stocks()
            if stocks.empty:
                st.info("수정할 stock이 없습니다.")
            else:
                sel = st.selectbox("Select stock_id", stocks["id"].tolist())
                row = stocks[stocks["id"] == sel].iloc[0].to_dict()
                with st.form("update_stock"):
                    st.write(f"**ID (immutable):** `{sel}`")
                    name = st.text_input("Reagent name *", value=row["name"])
                    stock_conc = st.number_input("Stock concentration *", min_value=0.0, value=float(row["stock_conc"]), step=0.1)
                    stock_unit = st.selectbox("Unit *", CONC_UNITS, index=CONC_UNITS.index(row["stock_unit"]) if row["stock_unit"] in CONC_UNITS else 0)
                    notes = st.text_input("Notes (optional)", value=row.get("notes") or "")
                    submitted = st.form_submit_button("Update")
                    if submitted:
                        try:
                            update_stock(conn, sel, name.strip(), stock_conc, stock_unit, notes.strip())
                            cached_load_stocks.clear()
                            st.success("Updated.")
                            st.rerun()
                        except psycopg2.IntegrityError as e:
                            st.error("업데이트 결과가 (name, conc, unit) 중복을 만들었습니다.")
                            st.caption(str(e))

        else:  # Delete
            stocks = cached_load_stocks()
            if stocks.empty:
                st.info("삭제할 stock이 없습니다.")
            else:
                sel = st.selectbox("Select stock_id", stocks["id"].tolist())
                st.warning("삭제하면 복구가 어렵습니다. (참조 중이면 제약으로 막힐 수 있음)")
                if st.button("Delete"):
                    try:
                        delete_stock(conn, sel)
                        st.success("Deleted.")
                        cached_load_stocks.clear()
                        st.rerun()
                    except psycopg2.IntegrityError as e:
                        st.error("다른 데이터에서 참조 중이라 삭제가 제한될 수 있습니다.")
                        st.caption(str(e))

        return
    # Template
    if page == "Templates":
        render_template_manager(conn)
        return
    # ---------------- New Reaction (Plan + Multiple reaction cards) ----------------
    if page == "New Reaction":
        st.subheader("반응 계획서")

        # ---- Plan UI ----
        st.markdown("### Plan")
        plan_title_in = st.text_input("Plan title", value="", key="plan_title_in")
        plan_notes_in = st.text_input("Plan notes (optional)", value="", key="plan_notes_in")

        if "current_plan_id" not in st.session_state:
            st.session_state["current_plan_id"] = None


        # ---- Common meta (applied to all reactions when saving) ----
        st.markdown("### Plan meta")
        m1, m2 = st.columns([1, 2])
        with m1:
            category = st.selectbox("Category", CATEGORIES, index=3, key="plan_category")
        with m2:
            title_prefix = st.text_input("Title prefix (optional)", value="", key="title_prefix")

        # ---- Load stocks once ----
        stocks = cached_load_stocks()
        stock_options = ["(DB 미등록: 임시 시약)"] + (stocks["id"].tolist() if not stocks.empty else [])

        # ---- Reaction cards state ----
        if "rx_cards" not in st.session_state:
            st.session_state["rx_cards"] = [new_reaction_card(0)]

        top1, top2, top3 = st.columns([1, 1, 2])

        if len(st.session_state["rx_cards"]) > 1 and top2.button("Remove last", key="rm_last_rx"):
            st.session_state["rx_cards"].pop()
            st.rerun()

        show_saved = top3.checkbox("Show saved reactions in this plan", value=True, key="show_saved_plan")

        st.divider()

        # ---- Templates list (shared list, but load applies per-card) ----
        tmpl_df = cached_load_templates()
        tmpl_names = ["(None)"] + (tmpl_df["name"].tolist() if not tmpl_df.empty else [])
        
        # ---- Render each reaction card ----
        all_card_results: List[Tuple[int, List[str], List[Dict[str, Any]], float, float]] = []  # (idx, errors, computed, total_uL, final_uL_total)

        for rx_idx, card in enumerate(st.session_state["rx_cards"]):
            rx_key = f"rx{rx_idx}"

            with st.container(border=True):
                # compact header row
                h1, h2, h3, h4, h5 = st.columns([2.2, 1, 1, 1, 1.2])
                card["name"] = h1.text_input("Reaction Name", value=card.get("name", f"Reaction {rx_idx+1}"), key=f"{rx_key}_name")

                if h2.button("Copy Reaction", key=f"{rx_key}_dup"):
                    st.session_state["rx_cards"].insert(rx_idx + 1, copy.deepcopy(card))
                    # rename the inserted card to avoid confusion
                    st.session_state["rx_cards"][rx_idx + 1]["name"] = f"{card['name']} (copy)"
                    st.rerun()

                if len(st.session_state["rx_cards"]) > 1 and h3.button("삭제", key=f"{rx_key}_del"):
                    st.session_state["rx_cards"].pop(rx_idx)
                    st.rerun()

                card["include_in_save_all"] = h4.checkbox("Include", value=bool(card.get("include_in_save_all", True)), key=f"{rx_key}_include")

                # final volume
                fv1, fv2, fv3 = st.columns([1.2, 1, 2.8])
                card["final_volume"] = fv1.number_input("Final volume", min_value=0.0, value=float(card.get("final_volume", 20.0)), step=1.0, key=f"{rx_key}_fv")
                card["final_vol_unit"] = fv2.selectbox("Unit", VOL_UNITS, index=VOL_UNITS.index(card.get("final_vol_unit", "uL")), key=f"{rx_key}_fvu")

                # template controls (per card)
                card["template_select"] = fv3.selectbox("Template", tmpl_names, index=tmpl_names.index(card.get("template_select", "(None)")) if card.get("template_select") in tmpl_names else 0, key=f"{rx_key}_tmpl_sel")

                tA, tB, tC, tD = st.columns([1, 1.2, 1.8, 3])
                load_btn = tA.button("Load", key=f"{rx_key}_tmpl_load")
                
                card["template_name_for_save"] = tC.text_input("Template name", value=card.get("template_name_for_save", ""), key=f"{rx_key}_tmpl_name")
                card["template_desc_for_save"] = tD.text_input("Template desc", value=card.get("template_desc_for_save", ""), key=f"{rx_key}_tmpl_desc")

                if load_btn and card["template_select"] != "(None)":
                    tmpl_row = tmpl_df.loc[
                    tmpl_df["name"] == card["template_select"]
                    ].iloc[0]

                    fv_key = f"{rx_key}_fv"
                    fvu_key = f"{rx_key}_fvu"

                    # 🔑 핵심: 기존 위젯 state 제거
                    if fv_key in st.session_state:
                        del st.session_state[fv_key]
                    if fvu_key in st.session_state:
                        del st.session_state[fvu_key]

                    # card에 template 값 세팅
                    card["final_volume"] = float(
                        tmpl_row["final_volume"]
                        if tmpl_row["final_volume"] is not None
                        else 20.0
                    )
                    card["final_vol_unit"] = str(
                        tmpl_row["final_volume_unit"]
                        if tmpl_row["final_volume_unit"] is not None
                        else "uL"
                    )

                    template_id = tmpl_row["template_id"]
                    items = load_template_items(None, template_id)

                    card["rows"] = [empty_row() for _ in range(len(items) if len(items) > 0 else 1)]

                    for i, row in items.reset_index(drop=True).iterrows():
                        r = card["rows"][i]
                        r["stock_sel"] = str(row["stock_id"])
                        r["target_conc"] = float(row["example_target"]) if row["example_target"] is not None else 0.0
                        r["target_unit"] = str(row["example_target_unit"] or "mM")
                        r["vol"] = float(row["example_volume"]) if row["example_volume"] is not None else 0.0
                        r["vol_unit"] = str(row["example_volume_unit"] or "uL")
                        r["note"] = str(row["item_note"] or "")

                    st.success(f"Template '{card['template_select']}' loaded")
                    st.rerun()


                # components count

                st.caption("각 row에서 Target conc 또는 Volume 중 하나만 채우세요. (Volume 직접 입력 row는 최대 1개)")
                st.markdown("### Components")

                if st.button("+ Add reagent", key=f"{rx_key}_add"):
                    card["rows"].append(empty_row())
                    st.rerun()

                for i, row in enumerate(card["rows"]):
                    with st.container(border=True):
                        c1, c2, c3, c4, c5, c6, c7 = st.columns([2.2, 1.2, 1, 1.2, 1, 2.2, 0.6])
                        rid = row["row_id"]
                        with c1:
                            row["stock_sel"] = st.selectbox(
                                f"Stock (row {i+1})",
                                stock_options,
                                index=(
                                    stock_options.index(row["stock_sel"])
                                    if row["stock_sel"] in stock_options
                                    else 0
                                ),
                                key=f"{rx_key}_stock_sel_{rid}"
                            )

                            if row["stock_sel"] == "(DB 미등록: 임시 시약)":
                                row["custom_name"] = st.text_input(
                                    "Custom reagent name *",
                                    row["custom_name"],
                                    key=f"{rx_key}_custom_name_{rid}"
                                )
                                sc1, sc2 = st.columns(2)
                                row["custom_stock_conc"] = sc1.number_input(
                                    "Stock conc *",
                                    value=row["custom_stock_conc"],
                                    key=f"{rx_key}_cconc_{rid}"
                                )
                                row["custom_stock_unit"] = sc2.selectbox(
                                    "Unit",
                                    CONC_UNITS,
                                    key=f"{rx_key}_cunit_{rid}"
                                )
                    
                        row["target_conc"] = c2.number_input(
                            "Target conc",
                            min_value=0.0,
                            step=0.0001,
                            format="%.4f",
                            key=f"{rx_key}_tc_{rid}"
                        )
                        tu_key = f"{rx_key}_tu_{rid}"

                        # 🔑 최초 생성 시 기본값 mM
                        if tu_key not in st.session_state:
                            st.session_state[tu_key] = "mM"

                        row["target_unit"] = c3.selectbox( # 기본 단위 mM로 고정
                            "Target unit",
                            CONC_UNITS,
                            key=tu_key
                        )

                        row["vol"] = c4.number_input("Volume", value=row["vol"], key=f"{rx_key}_vol_{rid}")
                        row["vol_unit"] = c5.selectbox("Vol unit", VOL_UNITS, key=f"{rx_key}_vu_{rid}")
                        row["note"] = c6.text_input("Note", row["note"], key=f"{rx_key}_note_{rid}")
                        
                        with c7:
                            if st.button("🗑", key=f"{rx_key}_del_row_{rid}"):
                                card["rows"].pop(i)
                                st.rerun()

                # compute + show result (per card)
                errors, computed, total_uL, final_uL_total = compute_reaction(card, stocks, stock_options)
                all_card_results.append((rx_idx, errors, computed, total_uL, final_uL_total))


                st.markdown("#### Result")
                if errors:
                    st.error("Validation failed:")
                    for e in errors:
                        st.write(f"- {e}")
                else:
                    if len(computed) == 0:
                        st.info("계산할 component가 없습니다. Target conc 또는 Volume을 입력하세요.")
                    else:
                        df = pd.DataFrame(computed).sort_values("line_no")
                        st.dataframe(
                            df[["line_no", "reagent", "stock_conc", "stock_unit", "target_conc", "target_unit",
                                "volume", "volume_unit", "amount", "amount_unit", "note"]],
                            width='stretch', hide_index=True
                        )
                        fill_uL = final_uL_total - total_uL
                        st.success(f"Fill remaining: {fill_uL:.3f} uL")

        st.divider()

        if st.button("➕ Add Reaction", width='stretch'):
            st.session_state["rx_cards"].append(
                new_reaction_card(len(st.session_state["rx_cards"]))
            )
            st.rerun()


        # ---- Save All ----
        st.markdown("### Save")
        save_col1, save_col2 = st.columns([2, 1])

        with save_col1:
            st.info("아래 버튼은 Include 체크된 Reaction들을 한 번에 저장합니다. (plan_id 아래에 여러 reaction 생성)")
        with save_col2:
            do_save_all = st.button("Save ALL included reactions to DB", key="save_all_btn")

        if do_save_all:
            # filter included and valid
            cards_to_save: List[Tuple[Dict[str, Any], List[Dict[str, Any]]]] = []
            problems: List[str] = []

            for rx_idx, errors, computed, _total_uL, _final_uL_total in all_card_results:
                card = st.session_state["rx_cards"][rx_idx]
                if not card.get("include_in_save_all", True):
                    continue
                if errors:
                    problems.append(f"{card.get('name','Reaction')} has validation errors.")
                    continue
                if len(computed) == 0:
                    problems.append(f"{card.get('name','Reaction')} has no computed rows.")
                    continue
                cards_to_save.append((card, computed))

            if problems:
                st.error("저장 전에 아래 문제를 해결하세요:")
                for p in problems:
                    st.write(f"- {p}")
            else:
                ok, msg, pid = save_all_to_db(
                    conn=conn,
                    plan_id=st.session_state.get("current_plan_id"),
                    plan_title=plan_title_in,
                    plan_notes=plan_notes_in,
                    category=category,
                    username=username,
                    title_prefix=title_prefix,
                    cards_to_save=cards_to_save,
                )
                if ok:
                    st.session_state["current_plan_id"] = pid
                    
                    cached_load_plans.clear()
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)
        st.markdown("### Export")

        if st.button("📤 Export reactions to Excel"):
            export_data = []

            for rx_idx, errors, computed, _total_uL, _final_uL_total in all_card_results:
                card = st.session_state["rx_cards"][rx_idx]
                if len(computed) == 0:
                    continue

                df = pd.DataFrame(computed)[
                    [
                        "line_no",
                        "reagent",
                        "stock_conc",
                        "stock_unit",
                        "target_conc",
                        "target_unit",
                        "volume",
                        "volume_unit",
                        "amount",
                        "amount_unit",
                        "note",
                    ]
                ]

                fill_uL = _final_uL_total - _total_uL

                if fill_uL > 1e-6:
                    df = pd.concat([
                        df,
                        pd.DataFrame([{
                            "line_no": "",
                            "reagent": "D.W.",
                            "stock_conc": "",
                            "stock_unit": "",
                            "target_conc": "",
                            "target_unit": "",
                            "volume": round(fill_uL, 3),
                            "volume_unit": "uL",
                            "amount": "",
                            "amount_unit": "",
                            "note": "Fill to final volume"
                        }])
                    ], ignore_index=True)


                export_data.append((card["name"], df))

            if not export_data:
                st.warning("Export할 reaction 결과가 없습니다.")
            else:
                excel_file = export_reactions_to_excel(export_data)
                # ---- Excel file name ----
                if st.session_state.get("current_plan_id"):
                    plan_title = plan_title_in.strip()
                else:
                    plan_title = ""

                if plan_title:
                    excel_name = f"{plan_title}.xlsx"
                else:
                    today = datetime.now().strftime("%Y%m%d")
                    excel_name = f"LabCalc_{today}_{username}.xlsx"

                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_file,
                    file_name=excel_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )


        # ---- Show saved in plan ----
        if show_saved and st.session_state.get("current_plan_id"):
            st.markdown("### Saved reactions in this plan")
            pid = st.session_state["current_plan_id"]
            saved = cached_reactions_in_plan(pid)

            st.dataframe(saved, width='stretch', hide_index=True)
    if page == "Plans":
        st.subheader("Reaction Plans Records")

        # ---- Load plans ----
        plans_df = cached_load_plans()


        if plans_df.empty:
            st.info("No saved plans found")
            return

        st.markdown("### Plans")
        plans_df["created_at"] = (
            pd.to_datetime(plans_df["created_at"]) + timedelta(hours=9)
        )

        plans_df = plans_df.rename(
            columns={"created_at": "created_at (KST)"}
        )

        st.dataframe(
            plans_df.drop(columns=["plan_id"]),
            width='stretch',
            hide_index=True,
            height = 300
        )
        plans_df["display_title"] = (
            plans_df["title"]
            + " · "
            + plans_df["created_at (KST)"].dt.strftime("%Y-%m-%d %H:%M")
        )


        # ---- Select plan ----
        query = st.text_input("Search plan title")
        if query:
            plans_df = plans_df[plans_df["title"].str.contains(query, case=False)]

        selected_plan = st.selectbox(
            "Select a plan",
            plans_df["display_title"].tolist()
        )



        plan_row = plans_df[
            plans_df["display_title"] == selected_plan
        ].iloc[0]

        plan_id = plan_row["plan_id"]
        

        st.markdown("### Plan info")
        st.write(f"- **Category:** {plan_row['category']}")
        st.write(f"- **Created by:** {plan_row['created_by']}")

        st.write(f"- **Created at:** {plan_row['created_at (KST)']}")

        if plan_row["notes"]:
            st.write(f"- **Notes:** {plan_row['notes']}")

        st.divider()

        # ---- Load reactions in plan ----
        rows = fetchall(
            """
            SELECT
                reaction_id,
                title,
                final_volume,
                final_volume_unit,
                created_at
            FROM reactions
            WHERE plan_id = %s
            ORDER BY created_at
            """,
            (plan_id,),
        )

        reactions_df = pd.DataFrame(
            rows,
            columns=[
                "reaction_id",
                "title",
                "final_volume",
                "final_volume_unit",
                "created_at",
            ],
        )

        st.markdown("### Reactions in this plan")

        if reactions_df.empty:
            st.info("No reactions in this plan")
            return
        reactions_df["created_at"] = (
            pd.to_datetime(reactions_df["created_at"]) + timedelta(hours=9)
        )

        reactions_df = reactions_df.rename(
            columns={"created_at": "created_at (KST)"}
        )

        st.dataframe(
            reactions_df.drop(columns=["reaction_id"]),
            width='stretch',
            hide_index=True
        )

        # ---- Select reaction ----
        selected_reaction = st.selectbox(
            "Select a reaction",
            reactions_df["title"].tolist()
        )

        reaction_row = reactions_df[
            reactions_df["title"] == selected_reaction
        ].iloc[0]

        reaction_id = reaction_row["reaction_id"]

        st.divider()

        # ---- Load reaction composition ----

        rows = fetchall(
            """
            SELECT
                ri.line_no,
                COALESCE(s.name, ri.custom_name) AS reagent,
                ri.stock_conc,
                ri.stock_unit,
                ri.target_conc,
                ri.target_conc_unit,
                ri.volume,
                ri.volume_unit,
                ri.amount,
                ri.amount_unit,
                ri.note
            FROM reaction_items ri
            LEFT JOIN stocks s ON ri.stock_id = s.id
            WHERE ri.reaction_id = %s
            ORDER BY ri.line_no
            """,
            (reaction_id,),
        )

        items_df = pd.DataFrame(
            rows,
            columns=[
                "line_no",
                "reagent",
                "stock_conc",
                "stock_unit",
                "target_conc",
                "target_conc_unit",
                "volume",
                "volume_unit",
                "amount",
                "amount_unit",
                "note",
            ],
        )

        st.markdown("### Reaction composition")
        st.dataframe(
            items_df,
            width='stretch',
            hide_index=True
        )

        st.markdown("### Export this plan")

        export_data = []

        for _, rx in reactions_df.iterrows():
            rx_id = rx["reaction_id"]

            rows = fetchall(
                """
                SELECT
                    ri.line_no,
                    COALESCE(s.name, ri.custom_name) AS reagent,
                    ri.stock_conc,
                    ri.stock_unit,
                    ri.target_conc,
                    ri.target_conc_unit,
                    ri.volume,
                    ri.volume_unit,
                    ri.amount,
                    ri.amount_unit,
                    ri.note
                FROM reaction_items ri
                LEFT JOIN stocks s ON ri.stock_id = s.id
                WHERE ri.reaction_id = %s
                ORDER BY ri.line_no
                """,
                (rx_id,),
            )

            df = pd.DataFrame(
                rows,
                columns=[
                    "line_no",
                    "reagent",
                    "stock_conc",
                    "stock_unit",
                    "target_conc",
                    "target_unit",
                    "volume",
                    "volume_unit",
                    "amount",
                    "amount_unit",
                    "note",
                ],
            )

            export_data.append((rx["title"], df))

        if export_data:
            excel_file = export_reactions_to_excel(export_data)

            st.download_button(
                label="📥 Export this plan (Excel)",
                data=excel_file,
                file_name=f"{plan_row['title']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )
        else:
            st.warning("No reactions available for export.")

    if e_page == "Labeling Efficiency":
        eps_df = cached_eps_db()
        eps_names = (
            eps_df["name"].dropna().unique().tolist()
        )

        st.subheader("Labeling Ratio (A = ε·c·l, l = 1 cm)")
        st.caption(
            "시약 이름을 입력하고 enter를 치면 DB에 저장된 정보가 로딩됩니다. DB에 없는 시약은 값을 직접 입력해주세요."
        )
        st.markdown("### Target")

        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            A_target = st.number_input(
                "A_target",
                min_value=0.0,
                step=0.001,
                format="%.3f"
            )

        with col2:
            target_wavelength = st.selectbox(
                "Target wavelength (nm)",
                [260, 280],
                index=0
             )
            target_unit = st.selectbox(
                "Target unit",
                ["µM", "nM", "mM"]
            )

        with col3:
            target_name = st.selectbox(
                "Target name",
                options=eps_names,
                index=None,
                placeholder="Protein / PNA / Adapter",
                key="target_name",
                accept_new_options=True
            )


            hit = eps_df[
                (eps_df["name"] == target_name)
                & (eps_df["wavelength"] == target_wavelength)
            ]

            if not hit.empty:
                eps_target_default = float(hit.iloc[0]["epsilon"])
            else:
                eps_target_default = 0.0


            if not hit.empty:
                eps_target_default = float(hit.iloc[0]["epsilon"])
                st.caption("ε loaded from DB")
            else:
                eps_target_default = 0.0
                st.caption("ε not found → manual input")


            eps_target = st.number_input(
                "ε_target (M⁻¹·cm⁻¹)",
                min_value=0.0,
                value=float(eps_target_default)
            )

            if st.button("Save ε (Target) to DB"):
                if target_name.strip() and eps_target > 0:
                    upsert_epsilon(
                        conn,
                        name=target_name.strip(),
                        wavelength=target_wavelength,
                        epsilon=float(eps_target),
                    )
                    cached_eps_db.clear()
                    st.success(f"ε saved for {target_name}")
                else:
                    st.error("Target name과 ε 값이 필요합니다.")

        st.divider()

        # Dye 입력
        st.markdown("### Dye")

        col4, col5, col6 = st.columns([1, 1, 2])
        
        # ---- Dye wavelength candidates from ε DB (미리 정의!) ----
        if "dye_name" in st.session_state and st.session_state["dye_name"]:
            dye_wavelength_options = (
                eps_df.loc[
                    eps_df["name"] == st.session_state["dye_name"],
                    "wavelength"
                ]
                .dropna()
                .astype(int)
                .unique()
                .tolist()
            )
        else:
            dye_wavelength_options = []


        with col4:
            A_dye = st.number_input(
                "A_dye",
                min_value=0.0,
                step=0.001,
                format="%.3f"
            )

        with col5:
            if dye_wavelength_options:
                dye_wavelength = st.selectbox(
                    "Dye λmax (nm)",
                    options=dye_wavelength_options,
                    index=0,
                    help="Automatically loaded from ε DB"
                )
            else:
                dye_wavelength = st.number_input(
                    "Dye λmax (nm)",
                    min_value=200,
                    max_value=800,
                    step=1,
                    value=646
                )

            dye_unit = st.selectbox(
                "Dye unit",
                ["µM", "nM", "mM"]
            )


        with col6:
            dye_name = st.selectbox(
                "Dye name",
                options=eps_names,
                index=None,
                placeholder="Cy3 / Cy5 / Alexa",
                key="dye_name",
                accept_new_options=True
            )
            # ---- Dye wavelength candidates from ε DB ----


            eps_dye_data = get_epsilon_value(
                conn,
                dye_name,
                dye_wavelength
            )

            if eps_dye_data:
                eps_dye_default = eps_dye_data["epsilon"]
                st.caption("ε loaded from DB")
            else:
                eps_dye_default = 0.0
                st.caption("ε not found → manual input")

            eps_dye = st.number_input(
                "ε_dye (M⁻¹·cm⁻¹)",
                min_value=0.0,
                value=float(eps_dye_default)
            )

            if st.button("Save ε (Dye) to DB"):
                if dye_name.strip() and eps_dye > 0:
                    upsert_epsilon(
                        conn,
                        name=dye_name.strip(),
                        wavelength=dye_wavelength,
                        epsilon=float(eps_dye),
                    )
                    cached_eps_db.clear()
                    st.success(f"ε saved for {dye_name}")
                else:
                    st.error("Dye name과 ε 값이 필요합니다.")

        st.divider()

        
        # 계산 (Beer–Lambert)
        # ---- correction factor 먼저 ----
        # ---- 초기값 선언 (UI/저장용) ----
        # ---- EtOH precipitation defaults (for DB save) ----
        initial_target_nmol: Optional[float] = None
        recovered_nmol: Optional[float] = None
        ethanol_efficiency: Optional[float] = None

        target_uM = None
        dye_uM = None
        labeling_ratio = None
        purity = None
        can_calc = False

        # correction factor (dye → target 파장)
        cf_df = cached_cf_db()
        cf = lookup_cf(cf_df, dye_name, target_wavelength)
        cf_used = cf if cf > 0 else None

        # Target concentration
        if A_target > 0 and eps_target > 0:
            A_target_corr = max(A_target - cf * A_dye, 0.0)
            target_uM = (A_target_corr / eps_target) * 1e6

        if cf > 0:
            st.caption(
                f"CF applied: A_target_corr = A_target − ({cf} × A_dye)"
            )
        else:
            st.caption("CF not applied (no correction factor found)")


        # Dye concentration
        if A_dye > 0 and eps_dye > 0:
            dye_uM = (A_dye / eps_dye) * 1e6

        # Labeling efficiency
        if target_uM is not None and dye_uM is not None:
            labeling_ratio = dye_uM / target_uM if target_uM > 0 else 0
            purity = (
                target_uM / (target_uM + dye_uM) * 100
                if (target_uM + dye_uM) > 0 else 0
            )
            can_calc = True
        st.markdown("### Calculated concentrations")

        c1, c2 = st.columns(2)
        with c1:
            st.metric("Target (µM)", f"{target_uM:.3f}" if target_uM else "–")
        with c2:
            st.metric("Dye (µM)", f"{dye_uM:.3f}" if dye_uM else "–")
        
        st.markdown("### Labeling efficiency")

        if labeling_ratio is not None:
            st.metric("Labeling ratio (dye / target)", f"{labeling_ratio:.3f}")
            st.metric("Chemical purity (%)", f"{purity:.1f}")
        else:
            st.caption("Target와 Dye가 모두 있을 때 계산됩니다.")


        st.divider()
        # precipitation recovery
        st.markdown("### Ethanol precipitation efficiency (Target only)")
        st.caption(
            "Recovery efficiency of target after ethanol precipitation. "
            "Based on UV–Vis concentration before and after precipitation."
        )
        # ---- Before precipitation ----
        st.markdown("**Before precipitation**")

        initial_target_nmol = st.number_input(
            "Initial target amount (nmol)",
            min_value=0.0,
            step=0.001,
            format="%.3f",
            help="Total amount of target added before ethanol precipitation"
        )

        # ---- After precipitation ----
        st.markdown("**After precipitation**")

        a1, a2 = st.columns(2)

        with a1:
            target_uM_after = st.number_input(
                "Target concentration after (µM)",
                min_value=0.0,
                step=0.001,
                format="%.3f"
            )

        with a2:
            V_after_uL = st.number_input(
                "Resuspension volume (µL)",
                min_value=0.0,
                step=1.0,
                value=20.0
            )
            # ---- EtOH precipitation calculation ----
            recovered_nmol = None
            ethanol_efficiency = None

            if (
                initial_target_nmol is not None
                and initial_target_nmol > 0
                and target_uM_after > 0
                and V_after_uL > 0
            ):
                # µM × µL = pmol → nmol
                recovered_nmol = target_uM_after * V_after_uL / 1000
                ethanol_efficiency = recovered_nmol / initial_target_nmol * 100

            # ---- EtOH result display ----
            if ethanol_efficiency is not None:
                c1, c2 = st.columns(2)
                with c1:
                    st.metric("Recovered target (nmol)", f"{recovered_nmol:.3f}")
                with c2:
                    st.metric("EtOH recovery (%)", f"{ethanol_efficiency:.3f}")
            else:
                st.caption("Enter initial amount and post-precipitation values to calculate recovery.")


        # purity
        st.divider()
        st.markdown("### UV purity (A260 / A280)")

        p1, p2 = st.columns(2)

        with p1:
            A260 = st.number_input(
                "A260",
                min_value=0.0,
                step=0.01
            )

        with p2:
            A280 = st.number_input(
                "A280",
                min_value=0.0,
                step=0.01
            )
        uv_purity: Optional[float] = None

        if A260 > 0 and A280 > 0:
            uv_purity = A260 / A280
        if uv_purity is not None:
            st.markdown(
                f"**UV purity (A260/A280):** `{uv_purity:.2f}`"
            )
        else:
            st.caption("A260과 A280을 입력하면 purity가 계산됩니다.")

        record_title = st.text_input(
            "Record title *",
            placeholder="e.g. Cy5–ProteinX labeling (batch 3)",
            key="labeling_record_title"
        )

        st.markdown("### Save labeling record")

        record_note = st.text_input(
            "Note (optional)",
            key="labeling_record_note"
        )


        if st.button("💾 Save labeling record"):
            if not can_calc:
                st.error("계산값이 없습니다. A, ε 값을 먼저 입력하세요.")
            else:
                record_id = str(uuid.uuid4())
                execute(
                    """
                    INSERT INTO labeling_records (
                        record_id, created_by,
                        title,
                        target_name, target_epsilon, A_target,
                        dye_name, dye_epsilon, A_dye,
                        target_uM, dye_uM,
                        labeling_ratio,
                        purity,
                        A260, A280, uv_purity,
                        cf_used,
                        etoh_initial_nmol,
                        etoh_recovered_nmol,
                        etoh_efficiency,
                        note
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,%s,
                        %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s,
                        %s
                    )
                    """,
                    (
                        record_id, username, record_title.strip(),
                        target_name, eps_target, A_target,
                        dye_name, eps_dye, A_dye,
                        target_uM, dye_uM,
                        labeling_ratio,
                        purity,
                        A260, A280, uv_purity,
                        cf_used,
                        initial_target_nmol,
                        recovered_nmol,
                        ethanol_efficiency,
                        record_note or None
                    )
                )

                cached_labeling_records.clear()
                st.success("✅ Labeling record saved.")
                

            # 저장
            st.divider()

            record_note = st.text_input("Note (optional)", key="label_note")

        
            st.divider()
            
    if e_page == "Labeling Records":
        st.subheader("Labeling Efficiency Records")

        df = cached_labeling_records()

        df["created_at"] = (
            pd.to_datetime(df["created_at"]) + timedelta(hours=9)
        )
        df = df.rename(columns={
            "title": "Title",
            "etoh_efficiency": "EtOH recovery (%)",
            "labeling_ratio": "Labeling ratio",
            "uv_purity": "UV purity (A260/A280)"
        })


        if df.empty:
            st.info("저장된 labeling record가 없습니다.")
        else:
            st.dataframe(df,width='stretch',hide_index=True,
                column_config={
                    "Title": st.column_config.TextColumn(
                        "Title", width="large"
                    )
                }
            )

    
    # ---------------- ε Database ----------------
    if e_page == "Epsilon DB":

        st.subheader("Epsilon (ε) Database")

        eps_df = cached_eps_db()
        cf_df = cached_cf_db()

        # 항상 먼저 정의 (스코프 에러 방지)
        eps_name_options = (
            eps_df["name"].dropna().unique().tolist()
            if not eps_df.empty else []
        )

        st.markdown("### Current ε values")
        if eps_df.empty:
            st.info("ε database is empty.")
        else:
            st.dataframe(
                eps_df.sort_values(["name", "wavelength"]),
                width="stretch",
                hide_index=True
            )

        st.divider()

        # ======================================================
        # ε MODE
        # ======================================================
        st.markdown("## ε Management")

        eps_mode = st.radio(
            "ε Mode",
            ["Add", "Update", "Delete"],
            horizontal=True,
            key="eps_mode_new"
        )

        # ---------------- ADD ----------------
        if eps_mode == "Add":

            c1, c2, c3 = st.columns(3)

            with c1:
                eps_name = st.text_input("Name *", key="eps_add_name")

            with c2:
                eps_wavelength = st.number_input(
                    "Wavelength (nm) *",
                    min_value=200,
                    max_value=800,
                    step=1,
                    value=650,
                    key="eps_add_wave"
                )

            with c3:
                eps_value = st.number_input(
                    "ε (M⁻¹·cm⁻¹) *",
                    min_value=0.0,
                    step=1000.0,
                    key="eps_add_val"
                )

            if st.button("Add ε"):
                if not eps_name.strip():
                    st.error("Name is required.")
                elif eps_value <= 0:
                    st.error("ε must be greater than 0.")
                else:
                    upsert_epsilon(
                        conn,
                        name=eps_name.strip(),
                        wavelength=int(eps_wavelength),
                        epsilon=float(eps_value),
                    )
                    cached_eps_db.clear()
                    st.success("ε added.")
                    st.rerun()

        # ---------------- UPDATE ----------------
        elif eps_mode == "Update":

            if eps_df.empty:
                st.info("No ε entries available.")
            else:
                eps_df_display = eps_df.copy()
                eps_df_display["label"] = (
                    eps_df_display["name"]
                    + " @ "
                    + eps_df_display["wavelength"].astype(str)
                    + " nm"
                )

                selected = st.selectbox(
                    "Select ε entry",
                    eps_df_display["label"]
                )

                row = eps_df_display[
                    eps_df_display["label"] == selected
                ].iloc[0]

                new_value = st.number_input(
                    "New ε value",
                    min_value=0.0,
                    value=float(row["epsilon"]),
                    step=1000.0
                )

                if st.button("Update ε"):
                    upsert_epsilon(
                        conn,
                        name=row["name"],
                        wavelength=int(row["wavelength"]),
                        epsilon=float(new_value),
                    )
                    cached_eps_db.clear()
                    st.success("ε updated.")
                    st.rerun()

        # ---------------- DELETE ----------------
        else:
            if eps_df.empty:
                st.info("No ε entries available.")
            else:
                eps_df_display = eps_df.copy()
                eps_df_display["label"] = (
                    eps_df_display["name"]
                    + " @ "
                    + eps_df_display["wavelength"].astype(str)
                    + " nm"
                )

                selected = st.selectbox(
                    "Select ε entry to delete",
                    eps_df_display["label"]
                )

                row = eps_df_display[
                    eps_df_display["label"] == selected
                ].iloc[0]

                st.warning("This action cannot be undone.")

                if st.button("Delete ε"):
                    delete_epsilon(
                        row["name"],
                        int(row["wavelength"])
                    )
                    cached_eps_db.clear()
                    st.success("ε deleted.")
                    st.rerun()

        st.divider()

        # ======================================================
        # CF MODE
        # ======================================================
        st.markdown("## Correction Factor (CF) Management")

        cf_mode = st.radio(
            "CF Mode",
            ["Add", "Update", "Delete"],
            horizontal=True,
            key="cf_mode_new"
        )

        # ---------------- ADD ----------------
        if cf_mode == "Add":

            c1, c2, c3 = st.columns(3)

            with c1:
                cf_dye = st.text_input("Dye name *", key="cf_add_name")

            with c2:
                cf_wave = st.selectbox(
                    "Target wavelength (nm) *",
                    [260, 280],
                    key="cf_add_wave"
                )

            with c3:
                cf_val = st.number_input(
                    "CF value *",
                    min_value=0.0,
                    step=0.001,
                    key="cf_add_val"
                )

            if st.button("Add CF"):
                if not cf_dye.strip():
                    st.error("Dye name is required.")
                elif cf_val <= 0:
                    st.error("CF must be greater than 0.")
                else:
                    upsert_cf(
                        conn,
                        dye_name=cf_dye.strip(),
                        target_wavelength=cf_wave,
                        factor=float(cf_val),
                        note="manual entry"
                    )
                    cached_cf_db.clear()
                    st.success("CF added.")
                    st.rerun()

        # ---------------- UPDATE ----------------
        elif cf_mode == "Update":

            if cf_df.empty:
                st.info("No CF entries available.")
            else:
                cf_df_display = cf_df.copy()
                cf_df_display["label"] = (
                    cf_df_display["dye_name"]
                    + " → "
                    + cf_df_display["target_wavelength"].astype(str)
                    + " nm"
                )

                selected = st.selectbox(
                    "Select CF entry",
                    cf_df_display["label"]
                )

                row = cf_df_display[
                    cf_df_display["label"] == selected
                ].iloc[0]

                new_val = st.number_input(
                    "New CF value",
                    min_value=0.0,
                    value=float(row["factor"]),
                    step=0.001
                )

                if st.button("Update CF"):
                    upsert_cf(
                        conn,
                        dye_name=row["dye_name"],
                        target_wavelength=int(row["target_wavelength"]),
                        factor=float(new_val),
                        note="updated"
                    )
                    cached_cf_db.clear()
                    st.success("CF updated.")
                    st.rerun()

        # ---------------- DELETE ----------------
        else:
            if cf_df.empty:
                st.info("No CF entries available.")
            else:
                cf_df_display = cf_df.copy()
                cf_df_display["label"] = (
                    cf_df_display["dye_name"]
                    + " → "
                    + cf_df_display["target_wavelength"].astype(str)
                    + " nm"
                )

                selected = st.selectbox(
                    "Select CF entry to delete",
                    cf_df_display["label"]
                )

                row = cf_df_display[
                    cf_df_display["label"] == selected
                ].iloc[0]

                st.warning("This action cannot be undone.")

                if st.button("Delete CF"):
                    delete_cf(
                        row["dye_name"],
                        int(row["target_wavelength"])
                    )
                    cached_cf_db.clear()
                    st.success("CF deleted.")
                    st.rerun()


if __name__ == "__main__":
    main()
